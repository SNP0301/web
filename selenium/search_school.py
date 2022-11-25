from selenium import webdriver
import time
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import pandas as pd
import openpyxl as op

## 00. Setup Environment
options = Options()
options.add_experimental_option('detach',True)

driver = webdriver.Chrome(options=options)
driver.get("https://schoolzone.emac.kr/gis/gis.do")

w_book = op.load_workbook('/Users/snp/Documents/schools.xlsx') ## input file directory
w_sheet = w_book.active

result_book = op.load_workbook('/Users/snp/Documents/result.xlsx') ## output file directory
result_sheet = result_book.active

## 01. Search for all data rows in input file (w_book)
i = 3
for i in range(1, 2800): 
    search_name = w_sheet.cell(row=i, column=3).value
    print(search_name)


    school_name = driver.find_element(By.ID,'searchText')
    school_name.send_keys(search_name + Keys.ENTER)

    e_frame = driver.find_element(By.XPATH,'//*[@id="searchIframe"]')
    driver.switch_to.frame(e_frame)

    try: ## 02. Get the element and axis data: x, y each
        elem = driver.find_element(By.XPATH,'/html/body/div[4]/div[1]/ul[1]')
        x_addr = elem.get_attribute('x')
        y_addr = elem.get_attribute('y')

        cmd_dup = "parent.fn_getSchoolArea(" + x_addr + ',' + y_addr + ',\'highSchoolArea\',\'vworld\');' ## highschool area
        ##cmd_dup = "parent.fn_getSchoolArea(" + x_addr + ',' + y_addr + ',\'middleSchoolArea\',\'vworld\');' ## middleschool area
        driver.execute_script(cmd_dup)
        driver.switch_to.default_content()

        ## 03. Move on search frame and extract school area text
        driver.find_element(By.XPATH,'/html/body/div[1]/div[10]/div[1]/p/input').send_keys(Keys.ENTER)
        i_frame = driver.find_element(By.XPATH,'/html/body/div[1]/div[2]/div[1]/div[2]/div/div[2]/iframe')
        driver.switch_to.frame(i_frame)

        area_name = driver.find_element(By.XPATH,'//*[@id="highSchoolArea"]/div/p').text ## highschool area
        ##area_name = driver.find_element(By.XPATH,'//*[@id="middleSchoolArea"]/div/p').text ## middleschool area

        result_sheet.cell(row=i, column=1).value = search_name
        result_sheet.cell(row=i, column=2).value = area_name
        result_book.save('/Users/snp/Documents/result.xlsx')
        print(area_name)

    except: ## if school area doesn't exist
        result_sheet.cell(row=i, column=1).value = search_name
        result_sheet.cell(row=i, column=2).value = "not exists"
        result_book.save('/Users/snp/Documents/result.xlsx')
        print('not exists')
       
    driver.switch_to.default_content()
    school_name.send_keys(Keys.COMMAND + 'a')
    school_name.send_keys(Keys.DELETE)

    time.sleep(0.15)

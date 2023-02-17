"""
23년 체크리스트 
1. 엑셀 파일 초기 위치 설정 (Line 28)
2. 엑셀 파일 저장 위치 설정 (Line 109)
3. 공시정보 조회시 년도를 22년에서 23년으로 변경 (Line 71)
"""

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC

import time
import openpyxl as op

def clear():
    search_box = driver.find_element(By.XPATH, '//*[@id="searchWord"]')
    search_box.click()
    search_box.send_keys(Keys.COMMAND + 'A')
    search_box.send_keys(Keys.DELETE)
    time.sleep(0.7)
    return True

### 1. 엑셀 파일 초기 위치 설정
output_book = op.load_workbook('/Users/snp/web/Selenium/output.xlsx')
### ㄴ 두 작은 따옴표 사이에 파일저장경로를 입력
###     ㄴ 윈도우의 경우, [C드라이브]-[Documents]-[Downloads]에 있는 엑셀.xlsx 파일에 저장하려면 아래와 같이 입력
###         ㄴ output_book = op.load_workbook('C:\Documents\Downloads\엑셀.xlsx')
output_sheet = output_book.active

### 학교 알리미 창 열기
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.schoolinfo.go.kr/ei/ss/pneiss_a03_s0.do#")
parent_window = driver.current_window_handle

### 학교 알리미 창의 학교 검색 입력칸 클릭 후 작업 시작
for i in range(2, 3296):
    time.sleep(0.5)
    clear()
    search_box = driver.find_element(By.XPATH, '//*[@id="searchWord"]')
    search_box.click()


    time.sleep(0.5)
    target_school_name = output_sheet.cell(row=i, column = 14).value
    ### ㄴ 엑셀 파일의 (i,14) 셀에 저장된 값을 검색할 학교 이름으로 선언
    clear()
    search_box.click()
    search_box.send_keys(target_school_name)

    time.sleep(0.5)
    school_name_title = '//*[@title=\"' + output_sheet.cell(row=i, column = 7).value + '\"]'
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,school_name_title)))
    driver.find_element(By.XPATH,school_name_title).click()

    ### 검색 버튼 클릭
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="webSearchButton"]')))
    driver.find_element(By.XPATH,'//*[@id="webSearchButton"]').click()
    all_windows = driver.window_handles
    child_window = [window for window in all_windows if window != parent_window][0]
    driver.switch_to.window(child_window)

    ### 공시년도 클릭
    WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="gsYear"]')))
    select = Select(driver.find_element(By.XPATH, '//*[@id="gsYear"]'))

    ### 2. 엑셀 파일 저장 위치 설정
    select.select_by_value('2022')
    ### ㄴ 2023년의 공시정보를 조회하려는 경우 아래와 같이 입력
        ### ㄴ select.select_by_value('2023')

    ### 선택버튼 클릭
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,'//*[@id="gsYearBtn"]')))
    driver.find_element(By.XPATH,'//*[@id="gsYearBtn"]').click()

    ### 학생현황 클릭
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,'/html/body/div/div[3]/div[2]/ul/li[3]/a')))
    driver.find_element(By.XPATH,'/html/body/div/div[3]/div[2]/ul/li[3]/a').click()


    ### 졸업생의 진로 현황 클릭 
    WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,'/html/body/div/div[3]/div[2]/div/div[3]/ul/li[7]/a')))
    driver.find_element(By.XPATH,'/html/body/div/div[3]/div[2]/div/div[3]/ul/li[7]/a').click()


    """
    추출하려는 열의 index 선언
    졸업자(1), 일반고(2), 특성화고(3), 과학고(4), 외고/국제고(5), 예고/체고(6)
    마이스터고(7), 자사고(9), 자공고(10), 기타(12), 취업(14), 미인정기관 진학(15), 무직자/미상 (16)
    """
    td_arr = [1,2,3,4,5,6,7,9,10,12,14,15,16]


    WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.XPATH,'/html/body/div/div[3]/div[4]/div[3]/div[1]/div[2]/div/div[1]/table/tbody/tr[1]/td[1]')))
    k = 15
    for j in range(0,13):
        xpath_parameter_male = '/html/body/div/div[3]/div[4]/div[3]/div[1]/div[2]/div/div[1]/table/tbody/tr[1]/td[' + str(td_arr[j]) + ']'
        output_sheet.cell(row=i,column=k).value = driver.find_element(By.XPATH,xpath_parameter_male).text


        xpath_parameter_female = '/html/body/div/div[3]/div[4]/div[3]/div[1]/div[2]/div/div[1]/table/tbody/tr[2]/td[' + str(td_arr[j]) + ']'
        output_sheet.cell(row=i,column=k+13).value = driver.find_element(By.XPATH,xpath_parameter_female).text

        k = k + 1
    ### 3. 공시정보 조회시 년도를 22년에서 23년으로 변경
    output_book.save('/Users/snp/web/Selenium/output.xlsx')

    driver.close()
    driver.switch_to.window(parent_window)
    time.sleep(0.5)

    clear()
    if(clear()==True):
        print("[%d]: %s done"%(i,target_school_name))
        time.sleep(0.3)
    else:
        time.sleep(0.3)
        clear()

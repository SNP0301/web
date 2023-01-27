from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select


import time
import pandas as pd
import openpyxl as op

### 0. 창을 연다 + 현재 창을 부모 창으로 지정
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get("https://www.schoolinfo.go.kr/ei/ss/pneiss_a03_s0.do#")
parent_window = driver.current_window_handle

output_book = op.load_workbook('/Users/snp/web/Selenium/output.xlsx') ## output file directory
output_sheet = output_book.active


### 1. [중학교] 클릭
middleschool_category = driver.find_element(By.XPATH, '//*[@id="mCSB_1_container"]/div/ul/li[2]/label')
middleschool_category.click()


### 2. [시, 도] 클릭
largearea_category = driver.find_element(By.XPATH, '//*[@id="mCSB_2_container"]/div/ul/li[1]/label')
output_sheet.cell(row=1, column=1).value = "1 by 1"
output_sheet.cell(row=1, column=2).value = largearea_category.text
output_sheet.cell(row=1, column=3).value = "hi"
largearea_category.click()
#### 서울특별시

### 3. [지역구] 클릭
smallarea_category = driver.find_element(By.XPATH,'//*[@id="level3Box"]/li[1]/label')
output_sheet.cell(row=1, column=3).value = smallarea_category.text
smallarea_category.click()

schools = driver.find_elements(By.XPATH, '//*[@id="level4Box"]');


#### 강남구

### 4. [중학교] 클릭
middleschool_name = driver.find_element(By.XPATH, '//*[@id="level4Box"]/li[1]/label')
output_sheet.cell(row=1, column=4).value = middleschool_name.text
middleschool_name.click()
#### 개원중학교

### 5. [검색] 클릭 (창 열림)
search_button = driver.find_element(By.XPATH, '//*[@id="webSearchButton"]')
search_button.click()

all_windows = driver.window_handles
child_window = [window for window in all_windows if window != parent_window][0]
driver.switch_to.window(child_window)

time.sleep(0.5)

### 6. 열린 창에서 공시정보 [22년] 선택, delay 0.05
select = Select(driver.find_element(By.XPATH, '//*[@id="gsYear"]'))
select.select_by_value('2022')
driver.find_element(By.XPATH,'//*[@id="gsYearBtn"]').click()
time.sleep(0.05)


### 7. [학생 현황] 선택
driver.find_element(By.XPATH,'/html/body/div/div[3]/div[2]/ul/li[3]/a').click()
time.sleep(0.05)

driver.find_element(By.XPATH,'/html/body/div/div[3]/div[2]/div/div[3]/ul/li[7]/a').click()
time.sleep(1)

num = driver.find_element(By.XPATH, '//*[@id="mCSB_11_container"]/table/tbody/tr[3]/td[1]').text

output_sheet.cell(row=1, column=5).value = num
output_book.save('/Users/snp/web/Selenium/output.xlsx')

time.sleep(500)
##driver.find_element(By.XPATH,'//*[@id="gsYear"]/option[value()="2022"]').click()




### 8. [졸업생의 진로 현황] 선택
### 9. 진학 및 진로 현황 스크랩


driver.close() ### 자녀 창 작업 종료
driver.switch_to.window(parent_window) ## 부모창으로 전환

middleschool_name = driver.find_element(By.XPATH, '//*[@id="level4Box"]/li[2]/label')
middleschool_name.click()
time.sleep(2) ### 개원중학교